import hashlib
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.bp_v6.excel_link import update_pdf_link


PDF_NAME = "逸飞AI智眼系统_创业大赛BP_V6.pdf"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _workbook(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "项目模板"
    first["A1"] = "项目名称"
    first["B2"] = "=1+1"
    first["J3"] = "旧版.pdf"
    first["J3"].hyperlink = "旧版.pdf"
    first["J3"].style = "Hyperlink"
    for index in range(2, 12):
        sheet = workbook.create_sheet(f"Sheet{index}")
        sheet["A1"] = f"保留-{index}"
    workbook.save(path)


def _snapshot(path: Path):
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            "sheets": tuple(workbook.sheetnames),
            "cells": {
                sheet.title: tuple(
                    (cell.coordinate, cell.value, cell.hyperlink.target if cell.hyperlink else None)
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None or cell.hyperlink is not None
                )
                for sheet in workbook.worksheets
            },
        }
    finally:
        workbook.close()


def test_copies_workbook_and_only_changes_first_sheet_pdf_link(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    _workbook(source)
    before_hash = _sha256(source)
    before = _snapshot(source)

    update_pdf_link(source, destination, PDF_NAME)

    assert _sha256(source) == before_hash
    after = _snapshot(destination)
    assert before["sheets"] == after["sheets"]
    assert after["cells"]["项目模板"][-1] == ("J3", PDF_NAME, PDF_NAME)
    assert before["cells"]["项目模板"][:-1] == after["cells"]["项目模板"][:-1]
    for sheet_name in before["sheets"][1:]:
        assert before["cells"][sheet_name] == after["cells"][sheet_name]


def test_rejects_non_relative_pdf_name(tmp_path):
    source = tmp_path / "source.xlsx"
    _workbook(source)
    try:
        update_pdf_link(source, tmp_path / "final.xlsx", "C:/absolute/file.pdf")
    except ValueError as error:
        assert "relative PDF filename" in str(error)
    else:
        raise AssertionError("absolute link should be rejected")
